'''
@File    :   train_cifar100_v3_1.py
@Time    :   2023/06/27 19:00
@Author  :   Hong Shijie
@Version :   3.1
@related :   cnn_dn_v2_2 backbone_v2_1 dn_torch_v7_2
@Note    :   在v3的基础上，
'''
import os

import torch.utils.data
import torchvision
from tqdm import tqdm
from torch import nn
from utils.concat.cnn_dn_v2_2 import concat_net
from utils.cnn.backbone_v1 import backbone
from torch.autograd import Variable
import pickle
import openpyxl
from openpyxl.styles import Alignment
import torch.distributed as dist

epochs = 25
batch_size = 32
dn_input_dim = [1, 512]
y_neuron_num = batch_size * 10
y_top_k = 1
y_bottom_up_percent = 0.5
z_neuron_num = 100
synapse_flag = False
dn_items = 1
learning_rate = 0.0001
datasets = 'cifar100'
version = 'test3d1'
save_path = '../results'
global_step = 0


def main():
    global global_step

    dist.init_process_group(backend="nccl")
    os.environ["CUDA_VISIBLE_DEVICES"] = "0, 1"

    train_dataset = torchvision.datasets.CIFAR100('../data/cifar100', train=True, download=True,
                                                  transform=torchvision.transforms.Compose([
                                                    torchvision.transforms.ToTensor(), ]))

    test_dataset = torchvision.datasets.CIFAR100('../data/cifar100', train=False, download=True,
                                                  transform=torchvision.transforms.Compose([
                                                    torchvision.transforms.ToTensor(), ]))

    train_sampler = torch.utils.data.distributed.DistributedSampler(train_dataset)
    test_sampler = torch.utils.data.distributed.DistributedSampler(test_dataset)

    train_loader = torch.utils.data.DataLoader(
        train_dataset,
        batch_size=batch_size,
        shuffle=True,
        num_workers=2
    )

    test_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR100('../data/cifar100', train=False, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=False)

    model = concat_net(batch_size, dn_input_dim, y_neuron_num, y_top_k, y_bottom_up_percent, z_neuron_num, synapse_flag)

    state_load = torch.load('../results/model_e246_acc726.pth')
    model.backbone.load_state_dict(state_load, strict=False)
    model.fc.weight.data = state_load['resnet18.fc.weight']
    model.fc.bias.data = state_load['resnet18.fc.bias']

    model = nn.DataParallel(model, device_ids=[0, 1])
    model = model.cuda()

    optimizer = torch.optim.SGD(model.module.backbone.parameters(), lr=learning_rate)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=10, gamma=0.5)

    save_dir = get_dir(save_path)

    for epo in range(epochs):

        unrevised_activated_num = train(train_loader, model, optimizer, scheduler, epo)
        revised_activated_num = remove(model, epo, save_dir)

        tra_dn_acc, tra_acc = test(train_loader, model, 'train')
        dn_acc, acc = test(test_loader, model, 'valid')

        torch.save(model.state_dict(), os.path.join(save_dir, 'model_e{}_acc{}.pth'.format(epo, int(acc * 100))))

        save_excel('mul', save_dir, epo, 0,
                   unrevised_num=unrevised_activated_num, revised_num=revised_activated_num,
                   train_dn_acc=tra_dn_acc, train_acc=tra_acc,
                   valid_dn_acc=dn_acc, valid_acc=acc)

        scheduler.step()
        global_step += 1


def remove(model_dn, epo, save_dir):
    if epo > 0:
        last_y_neuron_age = torch.zeros(model_dn.dn.y_neuron_age.data.size())
        temp_y_neuron_age = torch.load(os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo-1)))
        last_y_neuron_age[:temp_y_neuron_age.size()[0], :temp_y_neuron_age.size()[1]] = temp_y_neuron_age
        # last_z_neuron_age = torch.load(os.path.join(save_dir, 'z_neuron_age_e{}.pth'.format(epo-1))).cuda()
        # last_y_threshold = torch.load(os.path.join(save_dir, 'y_threshold_e{}.pth'.format(epo-1))).cuda()

        judge1 = (model_dn.dn.y_neuron_age.data.cpu() - last_y_neuron_age).squeeze(0)
        judge1 = torch.where(judge1 == 0, 1, 0)

        judge2 = torch.ones(model_dn.dn.y_neuron_age.data.size()).squeeze(0)
        judge2.index_put_(indices=tuple(model_dn.dn.max_index.unsqueeze(0).to(torch.int64)), values=torch.zeros(model_dn.dn.max_index.size()))

        judge3 = torch.where(model_dn.dn.y_neuron_age.data.cpu() == 0, 0, 1).squeeze(0)

        temp_y_indices = torch.nonzero(judge1 * judge2 * judge3 == 1).t()

        y_activated = torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()
        y_activated = int(torch.sum(y_activated))
        model_dn.dn.z_neuron_age.data = torch.floor(
            (1 - len(temp_y_indices.squeeze(0)) / y_activated) * model_dn.dn.z_neuron_age.data)

        model_dn.dn.y_neuron_age.data.index_put_(
            indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
            values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

        model_dn.dn.y_threshold.data.index_put_(
            indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
            values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

    torch.save(model_dn.dn.y_neuron_age.data.cpu(), os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo)))
    model_dn.dn.max_index = torch.empty(0)

    revised_activated_num = int(torch.sum(torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return revised_activated_num


def get_dir(saver_dir):
    cnt = 0
    while True:
        saver_f = os.path.join(saver_dir, '{}_v{}_'.format(datasets, version))
        if os.path.exists(saver_f + str(cnt)):
            cnt += 1
        else:
            break
    saver = saver_f + str(cnt)
    os.mkdir(saver)
    return saver


def save_excel(state, save_dir, epoch, start, **kwargs):
    cols = [chr(i) for i in range(65, 91)]
    for i in cols:
        if start + len(kwargs) >= len(cols):
            cols_add = ['{}{}'.format(i, chr(j)) for j in range(65, 91)]
            cols = cols + cols_add
        else:
            break

    file = os.path.join(save_dir, 'logs.xlsx')
    if os.path.exists(file):
        workbook = openpyxl.load_workbook(file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if epoch == 0:
        alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells('{}1:{}1'.format(cols[1 + start], cols[start + len(kwargs)]))
        sheet['{}1'.format(cols[1 + start])] = state
        sheet['{}1'.format(cols[1 + start])].alignment = alignment

        sheet.merge_cells('A1:A2')
        sheet['A1'] = 'Epoch'
        sheet['A1'].alignment = alignment

    for key, value in kwargs.items():
        if epoch == 0:
            sheet['{}2'.format(cols[1 + start])] = key

        sheet['{}{}'.format(cols[1 + start], str(3 + epoch))] = value
        start += 1

    sheet['A{}'.format(str(3 + epoch))] = epoch
    workbook.save(file)


def train(train_loader, model, optimizer, scheduler, epo):
    global global_step, y_neuron_num, batch_size

    loop_train = tqdm(train_loader, ncols=120)

    class_criterion = nn.CrossEntropyLoss()

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model.train()

    for img, label in loop_train:
        img, label = img.to(device), label.to(device)
        img, label = Variable(img), Variable(label, requires_grad=False)

        output, activated_num = model(img, label, 'train', dn_items, global_step)
        activated_num = torch.max(activated_num)

        loss = class_criterion(output, label)

        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        if (y_neuron_num - activated_num) < batch_size:
            y_neuron_num += batch_size * 100
            model.module.dn.reset(y_neuron_num)
            model.module.dn = model.module.dn.cuda()

        loop_train.set_description(f"Epoch: {epo}")
        loop_train.set_postfix(loss=loss.item(), lr=scheduler.get_last_lr()[0],
                               activated_number=activated_num.item(), memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))

    unrevised_activated_num = int(torch.sum(torch.where(model.module.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return unrevised_activated_num


def test(test_loader, model, name):
    loop_test = tqdm(test_loader, ncols=120)
    num_correct = 0
    fc_num_correct = 0
    dn_num_correct = 0

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model.eval()

    with torch.no_grad():
        for img, label in loop_test:
            img, label = img.to(device), label.to(device)

            output_dn, output = model(img, label, 'test', dn_items, global_step)

            num_correct += (output.argmax(1) == label).sum()
            acc = float(num_correct) / len(test_loader.dataset)

            # fc_num_correct += (output_fc.argmax(1) == label).sum()
            # fc_acc = float(fc_num_correct) / len(test_loader.dataset)

            dn_num_correct += (output_dn.argmax(1) == label).sum()
            dn_acc = float(dn_num_correct) / len(test_loader.dataset)

            loop_test.set_description("Test({}):".format(name))
            loop_test.set_postfix(acc=acc, dn_acc=dn_acc, memory='{}MB'.format(int(torch.cuda.memory_allocated() / 1024 / 1024)))
    return dn_acc, acc


if __name__ == '__main__':
    main()
