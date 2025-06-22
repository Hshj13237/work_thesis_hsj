'''
@File    :   train_cifar10_v2_5.py
@Time    :   2023/06/05 15:00
@Author  :   Hong Shijie
@Version :   2.5
@related :   cnn_dn_v1_3 backbone_v1 dn_torch_v6_8_1
@Note    :   在v2.4的基础上,可excel保存，并在每个batch中循环锁定（非epoch中）
'''
import os

import torch.utils.data
import torchvision
from tqdm import tqdm
from torch import nn
from utils.concat.cnn_dn_v1_3 import concat_net
from torch.autograd import Variable
import pickle
import openpyxl
from openpyxl.styles import Alignment

epochs = 25
batch_size = 32
dn_input_dim = [1, 512]
y_neuron_num = batch_size * 100
y_top_k = 1
y_bottom_up_percent = 0.5
z_neuron_num = 10
synapse_flag = False
dn_items = 1
learning_rate = 0.0001
datasets = 'cifar10'
version = '2d5'
save_path = '../results'
backbone_times = 3




def main():
    train_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR10('../data/cifar10', train=True, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=True)

    test_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR10('../data/cifar10', train=False, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=False)

    model = concat_net(batch_size, dn_input_dim, y_neuron_num, y_top_k, y_bottom_up_percent, z_neuron_num, synapse_flag)
    model = model.to(model.device)

    state_load = torch.load('../results/model_e60_acc92.pth')
    state_load.pop('resnet18.fc.weight')
    state_load.pop('resnet18.fc.bias')
    model.backbone.load_state_dict(state_load)

    optimizer_b = torch.optim.SGD(model.backbone.parameters(), lr=learning_rate)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer_b, step_size=10, gamma=0.5)

    optimizer_d = torch.optim.SGD(model.dn.parameters(), lr=learning_rate)

    save_dir = get_dir(save_path)

    for epo in range(epochs):
        unrevised_activated_num = train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo)

        revised_activated_num = remove(model, epo, save_dir)

        tra_acc = test(train_loader, model, 'train')
        acc = test(test_loader, model, 'valid')

        save_excel('联合bp', save_dir, epo, 0,
                   unrevised_num=unrevised_activated_num, revised_num=revised_activated_num,
                   train_acc=tra_acc, valid_acc=acc)

        torch.save(model.state_dict(), os.path.join(save_dir, 'model_e{}_acc{}.pth'.format(epo, int(acc * 100))))
        scheduler.step()


def save_excel(state, save_dir, epoch, start, **kwargs):
    cols = [chr(i) for i in range(65, 91)]
    for i in cols:
        if start + len(kwargs) >= len(cols):
            cols_add = ['{}{}'.format(i, chr(j)) for j in range(65, 91)]
            cols = cols + cols_add
        else:
            break

    file = os.path.join(save_dir, 'logs.xlsx')
    if os.path.exists(file):
        workbook = openpyxl.load_workbook(file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if epoch == 0:
        alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells('{}1:{}1'.format(cols[1 + start], cols[start + len(kwargs)]))
        sheet['{}1'.format(cols[1 + start])] = state
        sheet['{}1'.format(cols[1 + start])].alignment = alignment

        sheet.merge_cells('A1:A2')
        sheet['A1'] = 'Epoch'
        sheet['A1'].alignment = alignment

    for key, value in kwargs.items():
        if epoch == 0:
            sheet['{}2'.format(cols[1 + start])] = key

        sheet['{}{}'.format(cols[1 + start], str(3 + epoch))] = value
        start += 1

    sheet['A{}'.format(str(3 + epoch))] = epoch
    workbook.save(file)


def get_dir(saver_dir):
    cnt = 0
    while True:
        saver_f = os.path.join(saver_dir, '{}_v{}_'.format(datasets, version))
        if os.path.exists(saver_f + str(cnt)):
            cnt += 1
        else:
            break
    saver = saver_f + str(cnt)
    os.mkdir(saver)
    return saver


def train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo):
    global y_neuron_num, batch_size, backbone_times
    loop_train = tqdm(train_loader, ncols=150)

    criterion = nn.CrossEntropyLoss()

    model.train()

    for img, label in loop_train:
        img, label = img.to(model.device), label.to(model.device)
        img, label = Variable(img), Variable(label, requires_grad=False)

        model.eval()
        activated_num = model(img, label, 'lock_backbone', dn_items)

        if (y_neuron_num - activated_num) < batch_size:
            y_neuron_num += batch_size * 100
            model.dn.reset(y_neuron_num)
            model.dn = model.dn.cuda()

        model.train()
        for cnt in range(backbone_times):
            output = model(img, label, 'lock_dn', dn_items)

            loss = criterion(output, label)
            optimizer_b.zero_grad()
            optimizer_d.zero_grad()
            loss.backward()
            optimizer_b.step()

        loop_train.set_description(f"Epoch: {epo}")
        loop_train.set_postfix(loss=loss.item(), lr=scheduler.get_last_lr()[0], activated_number=activated_num.item(),
                               memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))

    unrevised_activated_num = int(torch.sum(torch.where(model.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return unrevised_activated_num


def remove(model_dn, epo, save_dir):
    if epo > 0:
        last_y_neuron_age = torch.zeros(model_dn.dn.y_neuron_age.data.size())
        temp_y_neuron_age = torch.load(os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo-1)))
        last_y_neuron_age[:temp_y_neuron_age.size()[0], :temp_y_neuron_age.size()[1]] = temp_y_neuron_age
        # last_z_neuron_age = torch.load(os.path.join(save_dir, 'z_neuron_age_e{}.pth'.format(epo-1))).cuda()
        # last_y_threshold = torch.load(os.path.join(save_dir, 'y_threshold_e{}.pth'.format(epo-1))).cuda()

        judge1 = (model_dn.dn.y_neuron_age.data.cpu() - last_y_neuron_age).squeeze(0)
        judge1 = torch.where(judge1 == 0, 1, 0)

        judge2 = torch.ones(model_dn.dn.y_neuron_age.data.size()).squeeze(0)
        judge2.index_put_(indices=tuple(model_dn.dn.max_index.unsqueeze(0).to(torch.int64)), values=torch.zeros(model_dn.dn.max_index.size()))

        judge3 = torch.where(model_dn.dn.y_neuron_age.data.cpu() == 0, 0, 1).squeeze(0)

        temp_y_indices = torch.nonzero(judge1 * judge2 * judge3 == 1).t()

        y_activated = torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()
        y_activated = int(torch.sum(y_activated))
        model_dn.dn.z_neuron_age.data = torch.floor(
            (1 - len(temp_y_indices.squeeze(0)) / y_activated) * model_dn.dn.z_neuron_age.data)

        model_dn.dn.y_neuron_age.data.index_put_(
            indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
            values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

        model_dn.dn.y_threshold.data.index_put_(
            indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
            values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

        model_dn.dn.y2z.weight.data[:, temp_y_indices.squeeze(0).numpy()] = 0

    torch.save(model_dn.dn.y_neuron_age.data.cpu(), os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo)))
    model_dn.dn.max_index = torch.empty(0)

    revised_activated_num = int(torch.sum(torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return revised_activated_num


def test(test_loader, model, name):
    loop_test = tqdm(test_loader, ncols=120)
    num_correct = 0
    total_loss = 0

    model.eval()

    with torch.no_grad():
        for img, label in loop_test:
            img, label = img.to(model.device), label.to(model.device)

            output = model(img, label, 'test', dn_items)
            num_correct += (output.argmax(1) == label).sum()

            acc = float(num_correct) / len(test_loader.dataset)
            loop_test.set_description("Test({}):".format(name))
            loop_test.set_postfix(acc=acc, memory='{}MB'.format(int(torch.cuda.memory_allocated() / 1024 / 1024)))

    return acc


if __name__ == '__main__':
    main()

