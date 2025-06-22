'''
@File    :   train_cifar100_v1_1.py
@Time    :   2023/07/25 15:00
@Author  :   Hong Shijie
@Version :   1.1
@related :   cnn_dn_v1_2/v3 backbone_v1_1/v3 dn_torch_v6_10/v6_11/v3
             cnn_dn_v3 backbone_v3 dn_torch_v6_11_1
@Note    :   在train_cifar1_v3.6的基础上,将数据集改成cifar100
'''
import os

import torch.utils.data
import torchvision
from tqdm import tqdm
from torch import nn
from utils.concat.cnn_dn_v3 import concat_net
from utils.cnn.backbone_v3 import ResNet18 as backbone
from torch.autograd import Variable
import torch.nn.functional as F
import pickle
import openpyxl
from openpyxl.styles import Alignment


epochs = 25
batch_size = 8
# dn_input_dim = [512, 1, 1]
dn_input_dim = [512, 4, 4]

y_neuron_num = batch_size * 100
y_top_k = 1
y_bottom_up_percent = 0.5
z_neuron_num = 100
synapse_flag = False
dn_items = 1
learning_rate = 0.0001
datasets = 'cifar100'
version = 'test1d1'
save_path = '../results'
backbone_times = 5

global_step = 0


def main():
    train_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR100('../data/cifar100', train=True, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=True)

    test_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR100('../data/cifar100', train=False, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=False)

    model_dn = concat_net(batch_size, dn_input_dim, y_neuron_num, y_top_k, y_bottom_up_percent, z_neuron_num,
                          synapse_flag)
    # model_dn = nn.DataParallel(model_dn).cuda()
    model_dn = model_dn.cuda()

    model_fc = backbone(z_neuron_num)
    # model_fc = nn.DataParallel(model_fc).cuda()
    model_fc = model_fc.cuda()

    # state_load = torch.load('../results/model_e246_acc726.pth')
    # key_mapping = {
    #     'resnet18.fc.weight': 'fc.weight',
    #     'resnet18.fc.bias': 'fc.bias',
    # }
    # state_load = {key_mapping.get(k, k): v for k, v in state_load.items()}
    # model_fc.load_state_dict(state_load)
    #
    # state_load.pop('fc.weight')
    # state_load.pop('fc.bias')
    # model_dn.backbone.load_state_dict(state_load)

    state_load = torch.load('../results/model_e473_acc76.pth')
    model_fc.load_state_dict(state_load)
    model_dn.backbone.load_state_dict(state_load)

    optimizer = torch.optim.SGD(model_fc.parameters(), lr=learning_rate)
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=15, gamma=0.5)

    save_dir = get_dir(save_path)

    for epo in range(epochs):
        unrevised_activated_num = train(train_loader, model_dn, model_fc, optimizer, scheduler, epo, 'dn', None)

        revised_activated_num = remove(model_dn, epo, save_dir)

        tra_fc_acc, tra_dn_acc = test(train_loader, model_dn, model_fc, 'train')
        fc_acc, dn_acc = test(test_loader, model_dn, model_fc, 'valid')

        save_excel('lock_backbone', save_dir, epo, 0,
                   unrevised_num=unrevised_activated_num, revised_num=revised_activated_num,
                   train_fc_acc=tra_fc_acc, train_dn_acc=tra_dn_acc,
                   valid_fc_acc=fc_acc, valid_dn_acc=dn_acc)

        torch.save(model_fc.state_dict(),
                   os.path.join(save_dir, 'model_tea_e{}_dn_acc{}.pth'.format(epo, int(fc_acc * 100))))
        torch.save(model_dn.state_dict(),
                   os.path.join(save_dir, 'model_stu_e{}_dn_acc{}.pth'.format(epo, int(dn_acc * 100))))

        for cnt in range(backbone_times):
            train(train_loader, model_dn, model_fc, optimizer, scheduler, epo, 'backbone', cnt)

            tra_fc_acc, tra_dn_acc = test(train_loader, model_dn, model_fc, 'train')
            fc_acc, dn_acc = test(test_loader, model_dn, model_fc, 'valid')

            save_excel('lock_dn({})'.format(cnt), save_dir, epo, 6 + cnt * 5, lr=scheduler.get_last_lr()[0],
                       train_fc_acc=tra_fc_acc, train_dn_acc=tra_dn_acc,
                       valid_fc_acc=fc_acc, valid_dn_acc=dn_acc)

            torch.save(model_fc.state_dict(),
                       os.path.join(save_dir, 'model_tea_e{}_bone{}_acc{}.pth'.format(epo, cnt, int(fc_acc * 100))))
            torch.save(model_dn.state_dict(),
                       os.path.join(save_dir, 'model_stu_e{}_bone{}_acc{}.pth'.format(epo, cnt, int(dn_acc * 100))))

        scheduler.step()


def remove(model_dn, epo, save_dir):
    if epo > 0:
        last_y_neuron_age = torch.zeros(model_dn.dn.y_neuron_age.data.size())
        temp_y_neuron_age = torch.load(os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo-1)))
        last_y_neuron_age[:temp_y_neuron_age.size()[0], :temp_y_neuron_age.size()[1]] = temp_y_neuron_age
        # last_z_neuron_age = torch.load(os.path.join(save_dir, 'z_neuron_age_e{}.pth'.format(epo-1))).cuda()
        # last_y_threshold = torch.load(os.path.join(save_dir, 'y_threshold_e{}.pth'.format(epo-1))).cuda()

        judge1 = (model_dn.dn.y_neuron_age.data.cpu() - last_y_neuron_age).squeeze(0)
        judge1 = torch.where(judge1 == 0, 1, 0)

        judge2 = torch.ones(model_dn.dn.y_neuron_age.data.size()).squeeze(0)
        judge2.index_put_(indices=tuple(model_dn.dn.max_index.unsqueeze(0).to(torch.int64)), values=torch.zeros(model_dn.dn.max_index.size()))

        judge3 = torch.where(model_dn.dn.y_neuron_age.data.cpu() == 0, 0, 1).squeeze(0)

        temp_y_indices = torch.nonzero(judge1 * judge2 * judge3 == 1).t()

        y_activated = torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()
        y_activated = int(torch.sum(y_activated))
        model_dn.dn.z_neuron_age.data = torch.floor(
            (1 - len(temp_y_indices.squeeze(0)) / y_activated) * model_dn.dn.z_neuron_age.data)

        model_dn.dn.y_neuron_age.data.index_put_(
            indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
            values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

        model_dn.dn.y_threshold.data.index_put_(
            indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
            values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

    torch.save(model_dn.dn.y_neuron_age.data.cpu(), os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo)))
    model_dn.dn.max_index = torch.empty(0)

    revised_activated_num = int(torch.sum(torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return revised_activated_num


def get_dir(saver_dir):
    cnt = 0
    while True:
        saver_f = os.path.join(saver_dir, '{}_v{}_'.format(datasets, version))
        if os.path.exists(saver_f + str(cnt)):
            cnt += 1
        else:
            break
    saver = saver_f + str(cnt)
    os.mkdir(saver)
    return saver


def save_excel(state, save_dir, epoch, start, **kwargs):
    cols = [chr(i) for i in range(65, 91)]
    for i in cols:
        if start + len(kwargs) >= len(cols):
            cols_add = ['{}{}'.format(i, chr(j)) for j in range(65, 91)]
            cols = cols + cols_add
        else:
            break

    file = os.path.join(save_dir, 'logs.xlsx')
    if os.path.exists(file):
        workbook = openpyxl.load_workbook(file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if epoch == 0:
        alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells('{}1:{}1'.format(cols[1 + start], cols[start + len(kwargs)]))
        sheet['{}1'.format(cols[1 + start])] = state
        sheet['{}1'.format(cols[1 + start])].alignment = alignment

        sheet.merge_cells('A1:A2')
        sheet['A1'] = 'Epoch'
        sheet['A1'].alignment = alignment

    for key, value in kwargs.items():
        if epoch == 0:
            sheet['{}2'.format(cols[1 + start])] = key

        sheet['{}{}'.format(cols[1 + start], str(3 + epoch))] = value
        start += 1

    sheet['A{}'.format(str(3 + epoch))] = epoch
    workbook.save(file)


def train(train_loader, model_dn, model_fc, optimizer, scheduler, epo, state, backbone_cnt):
    global global_step, y_neuron_num, batch_size

    loop_train = tqdm(train_loader, ncols=150)

    model_dn.eval()
    model_fc.train()

    if state == 'dn':
        with torch.no_grad():
            for img, label in loop_train:
                img, label = img.to(model_dn.device), label.to(model_dn.device)
                img, label = Variable(img), Variable(label, requires_grad=False)

                activated_num = model_dn(img, label, 'train', dn_items, global_step)

                optimizer.zero_grad()

                if (y_neuron_num - activated_num) < batch_size:
                    y_neuron_num += batch_size * 100
                    model_dn.dn.reset(y_neuron_num)
                    model_dn.dn = model_dn.dn.cuda()

                loop_train.set_description(f"Epoch: {epo}(lock_backbone)")
                loop_train.set_postfix(activated_number=activated_num.item(),
                                       memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))

        unrevised_activated_num = int(torch.sum(torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()))
        return unrevised_activated_num

    elif state == 'backbone':
        for img, label in loop_train:
            img, label = img.to(model_dn.device), label.to(model_dn.device)
            img, label = Variable(img), Variable(label, requires_grad=False)

            output_dn = model_dn(img, label, 'test', dn_items, global_step)
            dn_min, _ = torch.min(output_dn, dim=1)
            dn_max, _ = torch.max(output_dn, dim=1)
            output_dn = (output_dn - dn_min.unsqueeze(1)) / (dn_max.unsqueeze(1) - dn_min.unsqueeze(1))

            output_fc = model_fc(img)
            output_fc = F.softmax(output_fc, dim=1)

            loss = losses(output_fc, output_dn, label)

            optimizer.zero_grad()

            if loss is not None:
                loss.backward()
                optimizer.step()
            else:
                loss = torch.tensor([0])

            update_ema_variables(model_fc, model_dn, 0.999, global_step)

            loop_train.set_description(f"Epoch: {epo} (lock_dn({backbone_cnt}))")
            loop_train.set_postfix(loss=loss.item(), lr=scheduler.get_last_lr()[0],
                                   memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))
        global_step += 1


def losses(output_fc, output_dn, label):
    # class_criterion = nn.CrossEntropyLoss()
    consistency_criterion = nn.MSELoss()

    fc_correct = output_fc.argmax(1) == label
    dn_correct = output_dn.argmax(1) != label
    correct = fc_correct * dn_correct

    flag = torch.nonzero(correct == True).squeeze(1).cpu().numpy()

    if len(flag) == 0:
        return None
    else:
        temp_fc = output_fc[flag, :]
        temp_dn = output_dn[flag, :]
        consistency_loss = consistency_criterion(temp_fc, temp_dn)

        return consistency_loss


def update_ema_variables(model, ema_model, alpha, global_step):
    # Use the true average until the exponential average is more correct
    alpha = min(1 - 1 / (global_step + 1), alpha)

    for ema_param, param in zip(ema_model.backbone.parameters(), model.parameters()):
    # for (name1, ema_param), (name2, param) in zip(ema_model.backbone.named_parameters(), model.named_parameters()):
    #     print(name1, '\t', name2)
        ema_param.data.mul_(alpha).add_(1 - alpha, param.data)


def test(test_loader, model_dn, model_fc, name):
    loop_test = tqdm(test_loader, ncols=120)
    fc_num_correct = 0
    dn_num_correct = 0

    model_fc.eval()
    model_dn.eval()

    with torch.no_grad():
        for img, label in loop_test:
            img, label = img.to(model_dn.device), label.to(model_dn.device)

            fc_output = model_fc(img)
            dn_output = model_dn(img, label, 'test', dn_items, global_step)

            fc_num_correct += (fc_output.argmax(1) == label).sum()
            dn_num_correct += (dn_output.argmax(1) == label).sum()

            fc_acc = float(fc_num_correct) / len(test_loader.dataset)
            dn_acc = float(dn_num_correct) / len(test_loader.dataset)

            loop_test.set_description("\tTest({}):".format(name))
            loop_test.set_postfix(fc_acc=fc_acc, dn_acc=dn_acc, memory='{}MB'.format(int(torch.cuda.memory_allocated() / 1024 / 1024)))
    return fc_acc, dn_acc


if __name__ == '__main__':
    main()
