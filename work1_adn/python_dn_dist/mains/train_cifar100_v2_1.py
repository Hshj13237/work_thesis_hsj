'''
@File    :   train_cifar100_v2.py
@Time    :   2023/09/26
@Author  :   Hong Shijie
@Version :   2
@related :   cnn_dn_v1_3 backbone_v1 dn_torch_v6_8_1/v6_8_2
@Note    :   图像与特征图交替
'''
import os

import torch.utils.data
import torchvision
import torch.nn.functional as F
from tqdm import tqdm
from torch import nn
from utils.concat.cnn_dn_v2_2 import concat_net
from torch.autograd import Variable
import pickle
import openpyxl
from openpyxl.styles import Alignment
import random
# import matplotlib.pyplot as plt

import argparse
import torch.distributed as dist

epochs = 100000
batch_size = 128
dn_input_dim = [1, 512]
# y_neuron_num = batch_size * 10
y_neuron_num = 128
y_top_k = 1
y_bottom_up_percent = 0.5
z_neuron_num = 100
synapse_flag = False
dn_items = 1
learning_rate = 0.0001
datasets = 'Cifar100D0420'
version = '2'
save_path = '../results'
backbone_times = 1


class pre_loss_fnction(torch.autograd.Function):
    @staticmethod
    def forward(ctx, x):
        x = torch.where(x <= 0., 1e-2, 1e2)
        # x = torch.log(x)
        return x

    @staticmethod
    def backward(ctx, grad_output):
        return grad_output

def main():
    print(torch.initial_seed())

    train_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR100('../data/cifar100', train=True, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=True)

    test_loader = torch.utils.data.DataLoader(
        torchvision.datasets.CIFAR100('../data/cifar100', train=False, download=True,
                                     transform=torchvision.transforms.Compose([
                                         # torchvision.transforms.Grayscale(num_output_channels=1),
                                         torchvision.transforms.ToTensor(),
                                         # torchvision.transforms.Normalize(
                                         #     (0,), (1,))
                                         # (0.1307,), (0.3081,))
                                     ])),
        batch_size=batch_size, shuffle=False)

    model = concat_net(batch_size, dn_input_dim, y_neuron_num, y_top_k, y_bottom_up_percent, z_neuron_num, synapse_flag)
    state_load = torch.load('../results/model_e557_acc76.pth')
    # state_load.pop('fc.weight')
    # state_load.pop('fc.bias')
    for i in ['dn.y_neuron_age', 'dn.y_threshold', 'dn.x2y.weight', 'dn.z2y.weight', 'dn.y2z.weight']:
        state_load.pop(i)
    model.load_state_dict(state_load, strict=False)

    # state_load = torch.load('../results/model_e797_acc76.pth')
    # y_neuron_num = state_load['dn.y_neuron_age'].size()[1]
    # model = concat_net(batch_size, dn_input_dim, y_neuron_num, y_top_k, y_bottom_up_percent, z_neuron_num, synapse_flag)
    # model.load_state_dict(state_load, strict=False)

    optimizer_b = torch.optim.SGD(model.backbone.parameters(), lr=learning_rate)
    # optimizer_b = torch.optim.SGD(nn.ModuleList([model.backbone, model.fc]).parameters(), lr=learning_rate)

    scheduler = torch.optim.lr_scheduler.StepLR(optimizer_b, step_size=25, gamma=0.5)

    optimizer_d = torch.optim.SGD(model.dn.parameters(), lr=learning_rate)

    save_dir = get_dir(save_path)

    # acc = 0

    for epo in range(epochs):

        # unrevised_activated_num = train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo,
        #                                 "lock_dn")
        unrevised_activated_num = train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo,
                                        "lock_backbone")

        # if acc > 0.764:
        #     unrevised_activated_num = train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo,
        #                                     "lock_dn")
        # else:
        #     unrevised_activated_num = train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo,
        #                                     "lock_backbone")

        revised_activated_num = remove(model, epo, save_dir)

        tra_acc = test(train_loader, model, 'train', epo)
        acc = test(test_loader, model, 'valid', epo)

        save_excel('联合bp', save_dir, epo, 0,
                   unrevised_num=unrevised_activated_num, revised_num=revised_activated_num,
                   train_acc=tra_acc, valid_acc=acc)

        if (epo % 20 == 0):
            torch.save(model.state_dict(), os.path.join(save_dir, 'model_e{}_acc{}.pth'.format(epo, int(acc * 100))))
        # scheduler.step()
        # model.dn.z_neuron_age.data = (torch.zeros((1, z_neuron_num))).to('cuda:0')

        # model.dn.cos_sim = torch.empty(0)

        # if epo % 50 == 0 and epo != 0:
        #     model.dn.sim_thr += 0.05


def save_excel(state, save_dir, epoch, start, **kwargs):
    cols = [chr(i) for i in range(65, 91)]
    for i in cols:
        if start + len(kwargs) >= len(cols):
            cols_add = ['{}{}'.format(i, chr(j)) for j in range(65, 91)]
            cols = cols + cols_add
        else:
            break

    file = os.path.join(save_dir, 'logs.xlsx')
    if os.path.exists(file):
        workbook = openpyxl.load_workbook(file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if epoch == 0:
        alignment = Alignment(horizontal='center', vertical='center')

        sheet.merge_cells('{}1:{}1'.format(cols[1 + start], cols[start + len(kwargs)]))
        sheet['{}1'.format(cols[1 + start])] = state
        sheet['{}1'.format(cols[1 + start])].alignment = alignment

        sheet.merge_cells('A1:A2')
        sheet['A1'] = 'Epoch'
        sheet['A1'].alignment = alignment

    for key, value in kwargs.items():
        if epoch == 0:
            sheet['{}2'.format(cols[1 + start])] = key

        sheet['{}{}'.format(cols[1 + start], str(3 + epoch))] = value
        start += 1

    sheet['A{}'.format(str(3 + epoch))] = epoch
    workbook.save(file)


def get_dir(saver_dir):
    cnt = 0
    while True:
        saver_f = os.path.join(saver_dir, '{}_v{}_'.format(datasets, version))
        if os.path.exists(saver_f + str(cnt)):
            cnt += 1
        else:
            break
    saver = saver_f + str(cnt)
    os.mkdir(saver)
    return saver


def train(train_loader, model, optimizer_b, scheduler, optimizer_d, epo, mode):
    global y_neuron_num, batch_size, backbone_times
    loop_train = tqdm(train_loader, ncols=150)

    criterion = nn.CrossEntropyLoss()
    # criterion = nn.NLLLoss()
    pre_loss_fn = pre_loss_fnction.apply

    model.train()

    for img, label in loop_train:
        img, label = img.to('cuda:1'), label.to('cuda:1')
        img, label = Variable(img), Variable(label, requires_grad=False)

        model.train()

        if mode == "lock_backbone":

            activated_num, activated_num2 = model(img, label, 'lock_backbone', dn_items, epo)

            if (y_neuron_num - activated_num) < batch_size:
                y_neuron_num += batch_size
                model.dn.reset(y_neuron_num)
                model.dn = model.dn.cuda()

            if (model.dn2.y_neuron_num - activated_num2) < batch_size:
                model.dn2.reset(model.dn2.y_neuron_num + batch_size)
                model.dn2 = model.dn2.cuda()

            optimizer_b.zero_grad()
            optimizer_d.zero_grad()
            loop_train.set_description(f"DNEpoch: {epo}")
            loop_train.set_postfix(lr=scheduler.get_last_lr()[0],
                                   activated_number=activated_num2.item(),
                                   memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))

        elif mode == "lock_dn":
            for cnt in range(backbone_times):
                output = model(img, label, 'lock_dn', dn_items, epo)
                loss = criterion(output, label.to('cuda:0'))
                optimizer_b.zero_grad()
                optimizer_d.zero_grad()
                loss.backward()
                optimizer_b.step()
                loop_train.set_description(f"BBEpoch: {epo}")
                loop_train.set_postfix(loss=loss.item(), lr=scheduler.get_last_lr()[0],
                                       memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))

        # # model.train()
        # for cnt in range(backbone_times):
        #
        #     # def capture_grad(model, input, output):
        #     #     # global inputs, outputs
        #     #     model.its = input
        #     #     model.ops = output
        #     #
        #     # model.dn.x2y.register_backward_hook(capture_grad)
        #
        #     # def capture_tensor(model, input):
        #     #     # global inputs, outputs
        #     #     model.its = input
        #     #
        #     # model.backbone.resnet18.layer4.register_forward_hook(capture_tensor)
        #
        #     output = model(img, label, 'lock_dn', dn_items, epo)
        #
        #     # if (output.argmax(1) == label.to('cuda:0')).sum() != batch_size:
        #     #     print()
        #
        #     pre_output = pre_loss_fn(output)
        #
        #     loss = criterion(output, label.to('cuda:0'))
        #
        #     optimizer_b.zero_grad()
        #     optimizer_d.zero_grad()
        #     # loss.backward()
        #     # optimizer_b.step()
        #
        #     # if (output.argmax(1) == label.to('cuda:0')).sum() != batch_size:
        #     #     activated_num = model(img, label, 'lock_backbone', dn_items, epo)
        #
        # # with torch.no_grad():
        # #     activated_num = model(img, label, 'lock_backbone', dn_items, epo)
        # #     new_output = model(img, label, 'test', dn_items, epo)
        # #     save_output(save_dir, epo, temp_cnt, first_res=output.argmax(1).data.cpu().numpy()[0],
        # #                 sec_res=new_output.argmax(1).data.cpu().numpy()[0], label=label.data.cpu().numpy()[0])
        # #
        # # temp_cnt += 1
        #
        # loop_train.set_description(f"Epoch: {epo}")
        # loop_train.set_postfix(loss=loss.item(), lr=scheduler.get_last_lr()[0], activated_number=activated_num.item(),
        #                        memory='{}MB'.format(int(torch.cuda.max_memory_allocated() / 1024 / 1024)))

    unrevised_activated_num = int(torch.sum(torch.where(model.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return unrevised_activated_num


def remove(model_dn, epo, save_dir):
    # del_idx = torch.nonzero(torch.sum(model_dn.dn.y2z.weight.data == 0.0, dim=0) < (z_neuron_num - 1)).squeeze(1).cpu()
    del_idx = torch.empty(0, dtype=torch.int64)

    if epo > 0:
        last_y_neuron_age = torch.zeros(model_dn.dn.y_neuron_age.data.size())
        temp_y_neuron_age = torch.load(os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo-1)))
        last_y_neuron_age[:temp_y_neuron_age.size()[0], :temp_y_neuron_age.size()[1]] = temp_y_neuron_age
        # last_z_neuron_age = torch.load(os.path.join(save_dir, 'z_neuron_age_e{}.pth'.format(epo-1))).cuda()
        # last_y_threshold = torch.load(os.path.join(save_dir, 'y_threshold_e{}.pth'.format(epo-1))).cuda()

        judge1 = (model_dn.dn.y_neuron_age.data.cpu() - last_y_neuron_age).squeeze(0)
        judge1 = torch.where(judge1 == 0, 1, 0)

        judge2 = torch.ones(model_dn.dn.y_neuron_age.data.size()).squeeze(0)
        judge2.index_put_(indices=tuple(model_dn.dn.max_index.unsqueeze(0).to(torch.int64)), values=torch.zeros(model_dn.dn.max_index.size()))

        judge3 = torch.where(model_dn.dn.y_neuron_age.data.cpu() == 0, 0, 1).squeeze(0)

        exc_y_indices = torch.nonzero(judge1 * judge2 * judge3 == 1).squeeze(1)
    else:
        exc_y_indices = torch.empty(0, dtype=torch.int64)

    exc_y_indices = torch.empty(0, dtype=torch.int64)

    temp_y_indices = torch.cat((exc_y_indices, del_idx))
    temp_y_indices = torch.unique(temp_y_indices).unsqueeze(0)

    y_activated = torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()
    y_activated = int(torch.sum(y_activated))
    model_dn.dn.z_neuron_age.data = torch.floor(
        (1 - len(temp_y_indices.squeeze(0)) / y_activated) * model_dn.dn.z_neuron_age.data)

    model_dn.dn.y_neuron_age.data.index_put_(
        indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
        values=torch.zeros(temp_y_indices.squeeze(0).size()).cuda())

    model_dn.dn.y_threshold.data.index_put_(
        indices=tuple(torch.cat((torch.zeros(temp_y_indices.size()), temp_y_indices), dim=0).to(torch.int64).cuda()),
        values=(torch.zeros(temp_y_indices.squeeze(0).size())).cuda())

    model_dn.dn.y2z.weight.data[:, temp_y_indices.squeeze(0).numpy()] = 0
    model_dn.dn.x2y.weight.data[temp_y_indices.squeeze(0).numpy(), :] = torch.rand((
        len(temp_y_indices.squeeze(0)), model_dn.dn.x2y.weight.size()[1])).cuda()
    model_dn.dn.z2y.weight.data[temp_y_indices.squeeze(0).numpy(), :] = torch.rand((
        len(temp_y_indices.squeeze(0)), model_dn.dn.z2y.weight.size()[1])).cuda()

    if epo > 0:
        os.remove(os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo-1)))
    torch.save(model_dn.dn.y_neuron_age.data.cpu(), os.path.join(save_dir, 'y_neuron_age_e{}.pth'.format(epo)))
    model_dn.dn.max_index = torch.empty(0)
    # model_dn.dn.cos_sim = torch.empty(0)

    revised_activated_num = int(torch.sum(torch.where(model_dn.dn.y_neuron_age >= 1, 1, 0).cpu()))
    return revised_activated_num


def test(test_loader, model, name, epo):
    loop_test = tqdm(test_loader, ncols=120)
    num_correct = 0
    total_loss = 0
    temp_cnt = 0

    model.eval()

    with torch.no_grad():
        for img, label in loop_test:
            # img, label = img.to(model.device), label.to(model.device)
            img, label = img.to('cuda:1'), label.to('cuda:0')

            output = model(img, label, 'test', dn_items, epo)
            num_correct += (output.argmax(1) == label).sum()

            # output_dn, output_fc = model(img, label, 'test', dn_items, epo)
            # idx = torch.nonzero(output_dn.argmax(1) != output_fc.argmax(1).to('cuda:0')).squeeze(1)
            #
            # for i in idx:
            #     temp_cnt += 1
            #     num_correct += (output_fc[i].argmax(0).to('cuda:0') == label[i])
            # if temp_cnt == 0:
            #     acc = 0
            # else:
            #     acc = float(num_correct) / float(temp_cnt)

            # num_correct += (output_dn.argmax(1) == output_fc.argmax(1).to('cuda:0')).sum()


            acc = float(num_correct) / len(test_loader.dataset)
            loop_test.set_description("Test({}):".format(name))
            loop_test.set_postfix(acc=acc, memory='{}MB'.format(int(torch.cuda.memory_allocated() / 1024 / 1024)))

    return acc


def save_output(save_dir, epoch, cnt, **kwargs):
    cols = [chr(i) for i in range(65, 91)]
    for i in cols:
        if (1 + epoch * len(kwargs) >= len(cols)):
            cols_add = ['{}{}'.format(i, chr(j)) for j in range(65, 91)]
            cols = cols + cols_add
        else:
            break

    file = os.path.join(save_dir, 'act.xlsx')
    if os.path.exists(file):
        workbook = openpyxl.load_workbook(file)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if epoch == 0:
        sheet['A1'] = "img_cnt"
        sheet['A{}'.format(2+cnt)] = cnt

    start = 0
    for key, value in kwargs.items():
        if cnt == 0:
            sheet['{}1'.format(cols[1 + epoch * len(kwargs) + start])] = key

        sheet['{}{}'.format(cols[1 + epoch * len(kwargs) + start], 2+cnt)] = value
        start += 1

    workbook.save(file)


if __name__ == '__main__':
    main()

